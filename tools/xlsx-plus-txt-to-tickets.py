import json, re, sqlite3, os
import openpyxl

XLSX = "/Users/OnTrackRetail/Downloads/DR_Spike_Data3.xlsx"
TXT = "/Users/OnTrackRetail/Downloads/utn_results.txt"
OUT = "data/dr-spike-3"
os.makedirs(OUT, exist_ok=True)
db = sqlite3.connect("data/spike.db")

def nlc_crs(nlc):
    r = db.execute("SELECT crs FROM locations WHERE nlc=? AND crs IS NOT NULL", (nlc,)).fetchone()
    return r[0] if r else None
def members(nlc):
    return {r[0] for r in db.execute("SELECT crs FROM locations WHERE fare_group_nlc=? AND crs IS NOT NULL", (nlc,)).fetchall()}
def resolve(nlc, scans, want):
    d = nlc_crs(nlc)
    if d: return d
    mem = members(nlc)
    for s in scans:
        sc = s["Scan"]
        if sc.get("action_text") == "Accepted" and sc.get("scan_mode") == want:
            c = nlc_crs(sc.get("STATION") or "")
            if c in mem: return c
    for s in scans:
        c = nlc_crs(s["Scan"].get("STATION") or "")
        if c in mem: return c
    return None
def times(t):
    return re.findall(r"(\d{1,2}:\d{2})", t or "")

# --- parse scan txt into UTN -> payload ---
scan_by_utn = {}
cur = None
with open(TXT) as f:
    for line in f:
        m = re.match(r"=====\s*UTN:\s*(\S+)", line)
        if m:
            cur = m.group(1); continue
        s = line.strip()
        if cur and s.startswith("{"):
            try:
                scan_by_utn[cur.upper()] = json.loads(s)
            except Exception as e:
                print(f"  ! JSON parse failed for {cur}: {e}")
            cur = None
print(f"Parsed {len(scan_by_utn)} scan blocks from txt")

# --- read spreadsheet ---
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active
HDR = 5
cols = {ws.cell(row=HDR, column=c).value: c for c in range(1, ws.max_column+1)}

written, missing, seen = 0, [], set()
for r in range(HDR+1, ws.max_row+1):
    utn = ws.cell(row=r, column=cols["UTN"]).value
    if not utn: continue
    key = str(utn).upper()
    if key not in scan_by_utn:
        missing.append(utn); continue
    if key in seen: continue  # duplicate row for same UTN
    seen.add(key)
    payload = scan_by_utn[key]
    try:
        ticket = payload["Response"][0]["Tickets"][0]["Ticket"]
    except Exception:
        missing.append(utn); continue
    scans = ticket.get("Scans", [])
    route = str(ws.cell(row=r, column=cols["Route Code"]).value or "0").zfill(5)
    ticket["RouteCode"] = route
    coupons = {s["Scan"].get("coupon_type") for s in scans}
    is_single = "Single" in coupons and "Outward" not in coupons

    def leg_for(itin_text, uid_text, origin_nlc, dest_nlc, coupon):
        t = times(itin_text)
        if len(t) < 2 or (uid_text and "|" in str(uid_text)):
            return None
        o = resolve(origin_nlc, scans, "entry")
        d = resolve(dest_nlc, scans, "exit")
        if not o or not d:
            return None
        return {"coupon_type": coupon,
                "legs": [{"origin_crs": o, "destination_crs": d,
                          "scheduled_departure": t[0], "scheduled_arrival": t[1]}]}

    itin = []
    out = leg_for(ws.cell(row=r, column=cols["Outward Ticket Itinerary"]).value,
                  ws.cell(row=r, column=cols["Outward Ticket Service UID"]).value,
                  ticket["OriginNLC"], ticket["DestinationNLC"],
                  "Single" if is_single else "Outward")
    if out: itin.append(out)
    ret = leg_for(ws.cell(row=r, column=cols["Return Ticket Itinerary"]).value,
                  ws.cell(row=r, column=cols["Return Ticket Service UID"]).value,
                  ticket["DestinationNLC"], ticket["OriginNLC"], "Return")
    if ret: itin.append(ret)
    if itin: ticket["Itinerary"] = itin

    with open(os.path.join(OUT, f"{key}.json"), "w") as fo:
        json.dump(payload, fo)
    written += 1

print(f"Wrote {written} ticket files to {OUT}/")
if missing:
    print(f"{len(missing)} UTNs had no usable scan data: {missing}")
db.close()
