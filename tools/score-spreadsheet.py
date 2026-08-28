import json, os
import openpyxl

XLSX = "/Users/OnTrackRetail/Downloads/DR_Spike_Data3.xlsx"
OUT_XLSX = "/Users/OnTrackRetail/Downloads/DR_Spike_Data3_scored.xlsx"
RESULTS = "data/dr-spike-3/results.json"
HDR = 5

results = {r["utn"].upper(): r["coupons"] for r in json.load(open(RESULTS))}

def fmt(m):
    if m is None:
        return "--:--"
    m = int(m) % 1440
    return f"{m // 60:02d}:{m % 60:02d}"

def journey_desc(c):
    legs = c.get("predictedLegs") or []
    if not legs:
        return f"{c['coupon']}: journey not resolved"
    o, d = legs[0], legs[-1]
    via = " -> ".join(l["destinationCrs"] for l in legs[:-1])
    path = f"{o['originCrs']} " + (f"-> {via} " if via else "") + f"-> {d['destinationCrs']}"
    return f"{c['coupon']}: {path}, dep {fmt(o['scheduledDeparture'])}, arr {fmt(d['actualArrival'])}"

REASONS = {
    "NOT_DELAYED": "on time",
    "BELOW_THRESHOLD": None,  # filled with delay minutes
    "INVALID_TICKET_FOR_SERVICE": "ticket not valid on the service travelled",
    "RESTRICTION_NOT_VALID": "off-peak ticket used on a service in a restricted (peak) time band",
    "ROUTE_NOT_PERMITTED": "route not permitted (e.g. HS1 on a non-HS1 ticket)",
    "OUTSIDE_VALIDITY": "travelled outside the ticket's valid time",
    "SERVICE_UNRESOLVED": "could not identify the service from the scans (possibly non-Southeastern)",
    "NO_TRAVEL_EVIDENCE": "no usable travel scans",
    "NO_HSP_DATA_YET": "no HSP performance data for the travel date",
}

def why(c):
    r = c.get("reason")
    if r == "BELOW_THRESHOLD":
        return f"{c['coupon']}: only {c['delayMinutes']} min late (under 15)"
    return f"{c['coupon']}: {REASONS.get(r, r or 'not eligible')}"

def is_unresolved(c):
    return c.get("confidence") == "UNKNOWN" or c.get("reason") in ("SERVICE_UNRESOLVED", "NO_TRAVEL_EVIDENCE", "NO_HSP_DATA_YET")

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
cols = {ws.cell(row=HDR, column=c).value: c for c in range(1, ws.max_column + 1)}
C_ENT, C_JNY, C_DLY, C_CMP, C_WHY = (cols["Entitled to compensation"], cols["What journey did the customer take"],
                                     cols["Delay length"], cols["Compensation amount"], cols["Why weren't they delayed"])

counts = {"yes": 0, "no": 0, "review": 0}
for r in range(HDR + 1, ws.max_row + 1):
    utn = ws.cell(row=r, column=cols["UTN"]).value
    if not utn:
        continue
    coupons = results.get(str(utn).upper())

    # clear target cells first
    for c in (C_ENT, C_JNY, C_DLY, C_CMP, C_WHY):
        ws.cell(row=r, column=c).value = None

    if not coupons:
        ws.cell(row=r, column=C_ENT).value = "review"
        ws.cell(row=r, column=C_WHY).value = "No ticket/scan data found for this UTN (check the UTN)"
        counts["review"] += 1
        continue

    entitled = [c for c in coupons if c.get("entitled")]
    if entitled:
        ws.cell(row=r, column=C_ENT).value = "yes"
        ws.cell(row=r, column=C_JNY).value = " | ".join(journey_desc(c) for c in entitled)
        ws.cell(row=r, column=C_DLY).value = " | ".join(f"{c['coupon']} {c['delayMinutes']} min" for c in entitled)
        pence = sum(c.get("compensationPence") or 0 for c in entitled)
        ws.cell(row=r, column=C_CMP).value = f"£{pence / 100:.2f}"
        counts["yes"] += 1
    elif any(is_unresolved(c) for c in coupons):
        ws.cell(row=r, column=C_ENT).value = "review"
        ws.cell(row=r, column=C_WHY).value = "; ".join(why(c) for c in coupons)
        counts["review"] += 1
    else:
        ws.cell(row=r, column=C_ENT).value = "no"
        ws.cell(row=r, column=C_WHY).value = "; ".join(why(c) for c in coupons)
        counts["no"] += 1

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX}")
print("Row outcomes:", counts)
